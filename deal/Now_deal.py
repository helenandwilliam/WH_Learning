#code - UTF-8

## coding:utf-8

import openpyxl

Pre_data = openpyxl.load_workbook('6.7美团和平区 .xlsx')
Next_data = openpyxl.load_workbook('6.21美团和平区.xlsx')


work = openpyxl.Workbook()
table = work.active
table.title = 'standard_list'

Pre_List_name = Pre_data.sheetnames #
Pre_sheet = Pre_data.get_sheet_by_name(Pre_List_name[0])#获取表格
Next_List_name = Next_data.sheetnames #
Next_sheet = Next_data.get_sheet_by_name(Next_List_name[0])#获取表格

Pre_Row_Num = Pre_sheet.max_row
Pre_Col_Num = Pre_sheet.max_column
Next_Row_Num = Next_sheet.max_row
Next_Col_Num = Next_sheet.max_column

List3_Row_Num = 0
Information_Num = 0

table.cell(row= 1, column= 1).value = '标准表店铺（早）'
table.cell(row= 1, column= 2).value = '店铺地址'
table.cell(row= 1, column= 3).value= '团购信息名'
table.cell(row= 1, column= 4).value= '销售量'
table.cell(row= 1, column= 5).value = '售价'
table.cell(row= 1, column= 6).value = '店铺名（晚）'
table.cell(row= 1, column= 7).value = '店铺地址'
table.cell(row= 1, column= 8).value= '团购信息名'
table.cell(row= 1, column= 9).value= '销售量'
table.cell(row= 1, column= 10).value = '售价'
table.cell(row= 1, column= 11).value = '营业额差值'

Same_column = 0

for i in range (2 ,Pre_Row_Num + 1 ):
	for j in range (2 , Next_Row_Num + 1):
		if Pre_sheet.cell(row=i, column=1).value == Next_sheet.cell(row=j, column=1).value and\
		   Pre_sheet.cell(row=i, column=2).value == Next_sheet.cell(row=j, column=2).value:

			Same_column = j

			for k in range (0,31):
				if Pre_sheet.cell(row=i, column=3).value == Next_sheet.cell(row=j+k, column=3).value and \
						Pre_sheet.cell(row=i, column=5).value == Next_sheet.cell(row=j+k, column=5).value:

					table.cell(row=i, column=1).value = Pre_sheet.cell(row=i, column=1).value
					table.cell(row=i, column=2).value = Pre_sheet.cell(row=i, column=2).value
					table.cell(row=i, column=3).value = Pre_sheet.cell(row=i, column=3).value
					table.cell(row=i, column=4).value = Pre_sheet.cell(row=i, column=4).value
					table.cell(row=i, column=5).value = Pre_sheet.cell(row=i, column=5).value
					table.cell(row=i, column=6).value = Next_sheet.cell(row=j+k, column=1).value
					table.cell(row=i, column=7).value = Next_sheet.cell(row=j+k, column=2).value
					table.cell(row=i, column=8).value = Next_sheet.cell(row=j+k, column=3).value
					table.cell(row=i, column=9).value = Next_sheet.cell(row=j+k, column=4).value
					table.cell(row=i, column=10).value = Next_sheet.cell(row=j+k, column=5).value
					break
'''
				elif (k == 30):

					table.cell(row= i, column= 1).value = Pre_sheet.cell(row=i, column=1).value
					table.cell(row= i, column= 2).value = Pre_sheet.cell(row=i, column=2).value
					table.cell(row= i, column= 3).value = Pre_sheet.cell(row=i, column=3).value
					table.cell(row= i, column= 4).value = 0
					table.cell(row= i, column= 5).value = 0
					#print(table.cell(row= i, column= 1).value)
					table.cell(row= i, column= 6).value = Next_sheet.cell(row=j+k, column=1).value
					table.cell(row= i, column= 7).value = Next_sheet.cell(row=j+k, column=2).value
					table.cell(row= i, column=8).value = Next_sheet.cell(row=j+k, column=3).value
					table.cell(row= i, column=9).value = 0
					table.cell(row= i, column=10).value = 0

'''


for row1 in range (2,Pre_Row_Num + 1):
	for column1 in range (1, 11):
		if table.cell(row=row1, column=column1).value is None:
			table.cell(row=row1, column=column1).value=0


#for row1 in range (2,Pre_Row_Num + 1):
'''
    if table.cell(row=row1, column=1).value is None:
        table.cell(row=row1, column=1).value = 0
    if table.cell(row=row1, column=2).value is None:
        table.cell(row=row1, column=2).value = 0
    if table.cell(row=row1, column=3).value is None:
       table.cell(row=row1, column=3).value = 0
    if table.cell(row=row1, column=4).value is None:
       table.cell(row=row1, column=4).value = 0
    if table.cell(row=row1, column=5).value is None:
        table.cell(row=row1, column=5).value = 0
    if table.cell(row=row1, column=6).value is None:
        table.cell(row=row1, column=6).value = 0
    if table.cell(row=row1, column=7).value is None:
        table.cell(row=row1, column=7).value=0
    if table.cell(row=row1, column=8).value is None:
        table.cell(row=row1, column=8).value=0
    if table.cell(row=row1, column=9).value is None:
        table.cell(row=row1, column=9).value=0
    if table.cell(row=row1, column=10).value is None:
        table.cell(row=row1, column=10).value=0
'''

table.cell(row=row1, column=4).value = float(table.cell(row=row1, column=4).value)
print(table.cell(row=row1, column=4).value )
table.cell(row=row1, column=5).value = float(table.cell(row=row1, column=5).value)
table.cell(row=row1, column=9).value = float(table.cell(row=row1, column=9).value)
table.cell(row=row1, column=10).value = float(table.cell(row=row1, column=10).value)

Amount = table.cell(row=row1, column=9).value - table.cell(row=row1, column=4).value
print(Amount)
table.cell(row=row1, column=11).value = Amount * table.cell(row=row1, column=10).value
print(table.cell(row=row1, column=11).value)
work.save('new1.xlsx')