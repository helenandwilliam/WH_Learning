import openpyxl

data = openpyxl.load_workbook('new_list.xlsx')

List_name_data = data.sheetnames
sheet_data = data.get_sheet_by_name(List_name_data[0])

Sheet_Row_Num = sheet_data.max_row
Sheet_Columns_Num = sheet_data.max_column


# 新建工作组，建立名为sheet1的表格
work = openpyxl.Workbook()
table = work.active
table.title = 'sheet1'

table.cell(row= 1, column= 1).value = '店铺编号'
table.cell(row= 1, column= 2).value= '店铺名'
table.cell(row= 1, column= 3).value= '店铺地址'
table.cell(row= 1, column= 4).value= '总营业额'

Business_num=0# 店铺数量
Aount = 0  #一家店铺的总销售额

for i in range (2 ,Sheet_Row_Num+1):
    if sheet_data.cell(row=i , column=1).value!=sheet_data.cell(row=i+1 , column=1).value:
        #两相邻第一个单元格内容不相等，必不是相同店铺，店铺数量加一
        Aount = Aount + sheet_data.cell(row=i , column=10).value
        Business_num = Business_num + 1
        table.cell(row=Business_num + 1, column=1).value = Business_num
        table.cell(row=Business_num + 1, column=2).value = sheet_data.cell(row=i, column=1).value
        table.cell(row=Business_num + 1, column=3).value = sheet_data.cell(row=i, column=2).value
        table.cell(row=Business_num + 1, column=4).value = Aount
    else:
        Aount = sheet_data.cell(row= i, column=9).value + sheet_data.cell(row= i + 1, column=9).value

work.save('new_end_list.xlsx')
