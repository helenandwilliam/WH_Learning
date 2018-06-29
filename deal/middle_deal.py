## coding:utf-8

import openpyxl
import string

data = openpyxl.load_workbook('6.21美团和平区.xlsx')
db = openpyxl.load_workbook('standard_list.xlsx')

List_name_data = data.sheetnames
sheet_data = data.get_sheet_by_name(List_name_data[0])
# print(sheet_data)
List_name_db = db.sheetnames
sheet_db = db.get_sheet_by_name(List_name_db[0])
# print(List_name_db)
# print(sheet_db)
Sheet_Row_Num = sheet_data.max_row
Sheet_Columns_Num = sheet_data.max_column
# print(Sheet_Row_Num,Sheet_Columns_Num)

Sheet_db_Row_Num = sheet_db.max_row
Sheet_db_Columns_Num = sheet_db.max_column
# print(Sheet_db_Row_Num,Sheet_db_Columns_Num)

# 新建工作组，建立名为sheet1的表格
work = openpyxl.Workbook()
table = work.active
table.title = 'sheet1'

List3_Row_Num = 0
Information_Num = 0

table.cell(row= 1, column= 1).value = '标准表店铺（5.14）'
table.cell(row= 1, column= 2).value= '团购信息名'
table.cell(row= 1, column= 3).value= '销售量'
table.cell(row= 1, column= 4).value = '售价'
table.cell(row= 1, column= 5).value = '店铺名（5.23）'
table.cell(row= 1, column= 6).value= '团购信息名'
table.cell(row= 1, column= 7).value= '销售量'
table.cell(row= 1, column= 8).value = '售价'


for i in range(2, Sheet_db_Row_Num + 1):
    for j in range(2, Sheet_Row_Num + 1):
        if sheet_db.cell(row=i, column=2).value == sheet_data.cell(row=j, column=1).value and \
                sheet_db.cell(row=i, column=3).value == sheet_data.cell(row=j, column=2).value:
            Information_Num = sheet_db.cell(row=i, column=4).value
            #print("值为%d\n\r", Information_Num)
            List3_Row_Num = List3_Row_Num + Information_Num
            for k in range(0, Information_Num):
        	    table.cell(row=List3_Row_Num - ((Information_Num - 2) - k), column=5).value = sheet_data.cell(row=j + k, column=1).value
        	    table.cell(row=List3_Row_Num - ((Information_Num - 2) - k), column=6).value = sheet_data.cell(row=j + k, column=3).value
        	    table.cell(row=List3_Row_Num - ((Information_Num - 2) - k), column=7).value = sheet_data.cell(row=j + k, column=4).value
        	    table.cell(row=List3_Row_Num - ((Information_Num - 2) - k), column=8).value = sheet_data.cell(row=j + k, column=5).value
            break
        else:
            if j == Sheet_Row_Num:
                Information_Num = sheet_db.cell(row=i, column=4).value
                List3_Row_Num = List3_Row_Num + Information_Num
                for k in range(0, Information_Num):
                    table.cell(row=List3_Row_Num - ((Information_Num - 2) - k), column=5).value = 0
                    table.cell(row=List3_Row_Num - ((Information_Num - 2) - k), column=6).value = 0
                    table.cell(row=List3_Row_Num - ((Information_Num - 2) - k), column=7).value = 0
                    table.cell(row=List3_Row_Num - ((Information_Num - 2) - k), column=8).value = 0
                break

    # else:




Standard_data = openpyxl.load_workbook('6.7美团和平区 .xlsx')

Standard_data_name = Standard_data.sheetnames
Standard_data_sheet_name= Standard_data .get_sheet_by_name(Standard_data_name[0])

Standard_data_Row_Num = Standard_data_sheet_name.max_row
Sheet_data_Columns_Num = Standard_data_sheet_name.max_column


for r in range(2,Standard_data_Row_Num + 1):
    #for c in range(1,4):
    table.cell(row=r, column=1).value = Standard_data_sheet_name.cell(row=r, column=1).value
    table.cell(row=r, column=2).value = Standard_data_sheet_name.cell(row=r, column=3).value
    table.cell(row=r, column=3).value = Standard_data_sheet_name.cell(row=r, column=4).value
    table.cell(row=r, column=4).value = Standard_data_sheet_name.cell(row=r, column=5).value


for row1 in range (2,Standard_data_Row_Num + 1):
    for column1 in range (2, 9):
        if table.cell(row=row1, column=column1).value == '':
             table.cell(row=row1, column=column1).value=0
for row1 in range (2,Standard_data_Row_Num + 1):
    if table.cell(row=row1, column=4).value is None:
        table.cell(row=row1, column=4).value = 0
    if table.cell(row=row1, column=8).value is None:
        table.cell(row=row1, column=8).value = 0

    table.cell(row=row1, column=3).value = float(table.cell(row=row1, column=3).value)
    table.cell(row=row1, column=4).value = float(table.cell(row=row1, column=4).value)
    table.cell(row=row1, column=7).value = float(table.cell(row=row1, column=7).value)
    table.cell(row=row1, column=8).value = float(table.cell(row=row1, column=8).value)

    Amount = table.cell(row=row1, column=7).value - table.cell(row=row1, column=3).value
    table.cell(row=row1, column=9).value = Amount * table.cell(row=row1, column=8).value

work.save('Mid_Deal_Table1.xlsx')


