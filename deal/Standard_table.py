## coding:utf-8

import openpyxl

data = openpyxl.load_workbook('6.7美团和平区 .xlsx')
#获取第一个表格 方法1
List_name = data.sheetnames
print(List_name)
#方法2
a = data.get_sheet_names()
print(a)

sheet = data.get_sheet_by_name(List_name[0])

Sheet_Row_Num = sheet.max_row
Sheet_Columns_Num = sheet.max_column

#新建工作组，建立名为standard_list的表格
work = openpyxl.Workbook()
table = work.active
#table.title = 'standard_list'

#for i in range(2, Sheet_Row_Num+1):
#	for j in range(1 , Sheet_Columns_Num):
#		if
table.cell(row= 1, column= 1).value = '店铺编号'
table.cell(row= 1, column= 2).value= '店铺名'
table.cell(row= 1, column= 3).value= '地址'
table.cell(row= 1, column= 4).value = '团购信息数目'

Business_num=0# 店铺数量
Goods_num=1   #一家店铺团购信息数量

for i in range (2 ,Sheet_Row_Num+1):
	if sheet.cell(row=i , column=1).value!=sheet.cell(row=i+1 , column=1).value:
		#两相邻第一个单元格内容不相等，必不是相同店铺，店铺数量加一
		Business_num = Business_num + 1

		table.cell(row=Business_num + 1, column=1).value = Business_num
		table.cell(row=Business_num + 1, column=2).value = sheet.cell(row=i, column=1).value
		table.cell(row=Business_num + 1, column=3).value = sheet.cell(row=i, column=2).value
		table.cell(row=Business_num + 1, column=4).value = Goods_num
		Goods_num = 1
	else:
		Goods_num = Goods_num + 1

work.save('standard_list.xlsx')
